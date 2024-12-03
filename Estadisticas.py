# Estadisticas.py

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os
import json
import pandas as pd
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, Table, TableStyle
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from tkcalendar import DateEntry
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

HISTORY_FILE = 'history.json'  # Asegúrate de que esta ruta sea correcta o ajusta según sea necesario

class Estadisticas:
    def __init__(self, master, parent):
        self.master = master
        self.parent = parent
        self.master.title("📊 Estadísticas de Uso - Algebrify")  # Título mejorado
        self.master.configure(bg=self.parent.bg_color if self.parent else 'white')
        self.master.state('zoomed')  # Maximizar ventana
        self.master.resizable(False, False)

        # Permitir salir del modo de pantalla completa con la tecla Escape
        self.master.bind("<Escape>", self.exit_fullscreen)

        # Definir estilos personalizados para los botones
        style = ttk.Style()
        style.theme_use('clam')  # Puedes elegir otros temas como 'default', 'alt', 'classic', 'vista'

        # Estilo para botones de exportación (Color Rojo Moderno)
        style.configure('ExportButton.TButton',
                        font=('Helvetica', 12, 'bold'),
                        foreground='white',
                        background='#E74C3C',  # Rojo moderno
                        padding=10)
        style.map('ExportButton.TButton',
                  background=[('active', '#C0392B')],
                  foreground=[('active', 'white')])

        # Estilo para botones generales (Color Azul Moderno)
        style.configure('GeneralButton.TButton',
                        font=('Helvetica', 12),
                        foreground='white',
                        background='#3498DB',  # Azul moderno
                        padding=10)
        style.map('GeneralButton.TButton',
                  background=[('active', '#2980B9')],
                  foreground=[('active', 'white')])

        # Estilo para Combobox
        style.configure('TCombobox',
                        foreground='black',
                        background='white',
                        fieldbackground='white',
                        font=('Helvetica', 12))

        # Crear un marco principal para organizar los elementos
        main_frame = ttk.Frame(self.master)
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)

        # Título de la sección de estadísticas
        title_label = ttk.Label(main_frame, text="📊 Estadísticas de Uso", font=("Helvetica", 24, "bold"))
        title_label.pack(pady=(0, 20))

        # Crear un Notebook para organizar las pestañas
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.pack(fill='both', expand=True)

        # Pestaña de Filtros
        self.filters_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.filters_tab, text='🔍 Filtros')

        # Pestaña de Resumen
        self.summary_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.summary_tab, text='📈 Resumen')

        # Pestaña de Gráficas
        self.charts_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.charts_tab, text='📊 Gráficas')

        # Pestaña de Exportación
        self.export_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.export_tab, text='💾 Exportar')

        # Inicializar filtros
        self.init_filters_tab()

        # Inicializar resumen
        self.init_summary_tab()

        # Inicializar gráficas
        self.init_charts_tab()

        # Inicializar exportación
        self.init_export_tab()

    def exit_fullscreen(self, event=None):
        self.master.state('normal')

    def init_filters_tab(self):
        """Inicializa los widgets de la pestaña de filtros."""
        # Frame para filtros
        filters_frame = ttk.Frame(self.filters_tab)
        filters_frame.pack(pady=20, padx=20, fill='x')

        # Filtro por Tipo de Ejercicio
        type_label = ttk.Label(filters_frame, text="Tipo de Ejercicio:", font=("Helvetica", 14))
        type_label.grid(row=0, column=0, sticky='w', pady=5, padx=5)

        self.type_var = tk.StringVar()
        self.type_combobox = ttk.Combobox(filters_frame, textvariable=self.type_var, state='readonly', style='TCombobox')
        self.type_combobox['values'] = self.get_all_types()
        self.type_combobox.current(0)  # Seleccionar 'Todos' por defecto
        self.type_combobox.grid(row=0, column=1, sticky='w', pady=5, padx=5)

        # Filtro por Rango de Fechas
        date_label = ttk.Label(filters_frame, text="Rango de Fechas:", font=("Helvetica", 14))
        date_label.grid(row=1, column=0, sticky='w', pady=5, padx=5)

        date_frame = ttk.Frame(filters_frame)
        date_frame.grid(row=1, column=1, sticky='w', pady=5, padx=5)

        start_label = ttk.Label(date_frame, text="Desde:", font=("Helvetica", 12))
        start_label.pack(side='left', padx=(0,5))

        self.start_date = DateEntry(date_frame, width=12, background='darkblue',
                                    foreground='white', borderwidth=2, date_pattern='y-mm-dd')
        self.start_date.pack(side='left', padx=(0,15))

        end_label = ttk.Label(date_frame, text="Hasta:", font=("Helvetica", 12))
        end_label.pack(side='left', padx=(0,5))

        self.end_date = DateEntry(date_frame, width=12, background='darkblue',
                                  foreground='white', borderwidth=2, date_pattern='y-mm-dd')
        self.end_date.pack(side='left', padx=(0,15))

        # Botón para aplicar filtros
        apply_button = ttk.Button(filters_frame, text="Aplicar Filtros", command=self.apply_filters, style='GeneralButton.TButton')
        apply_button.grid(row=2, column=0, columnspan=2, pady=20, padx=5, sticky='ew')

    def get_all_types(self):
        """Obtiene una lista de todos los tipos de ejercicios presentes en el historial."""
        history = self.read_history()
        types = set()
        for exercise in history:
            types.add(exercise.get('type', 'Desconocido'))
        types = sorted(list(types))
        types.insert(0, "Todos")  # Añadir opción para seleccionar todos
        return types

    def apply_filters(self):
        """Aplica los filtros seleccionados y actualiza las estadísticas y gráficas."""
        self.filtered_history = self.filter_history()
        if not self.filtered_history:
            messagebox.showinfo("Filtrar", "No se encontraron ejercicios que coincidan con los filtros.")
        # Actualizar resumen y gráficas
        self.update_summary_tab()
        self.update_charts_tab()

    def filter_history(self):
        """Filtra el historial según los filtros seleccionados."""
        history = self.read_history()
        selected_type = self.type_var.get()
        start_date = self.start_date.get_date()
        end_date = self.end_date.get_date()

        filtered = []
        for exercise in history:
            # Filtrar por tipo
            if selected_type != "Todos" and exercise.get('type') != selected_type:
                continue
            # Filtrar por fecha
            exercise_date_str = exercise['details'].get('date', '')
            try:
                exercise_date = datetime.strptime(exercise_date_str, '%Y-%m-%d').date()
            except ValueError:
                # Si la fecha no está en el formato esperado, omitir el ejercicio
                continue
            if start_date <= exercise_date <= end_date:
                filtered.append(exercise)
        return filtered

    def init_summary_tab(self):
        """Inicializa los widgets de la pestaña de resumen."""
        self.summary_frame = ttk.Frame(self.summary_tab)
        self.summary_frame.pack(pady=20, padx=20, fill='both', expand=True)

        # Estadísticas Clave
        self.total_label = ttk.Label(self.summary_frame, text="Total de Ejercicios Resueltos: 0", font=("Helvetica", 14))
        self.total_label.pack(anchor='w', pady=5)

        self.average_label = ttk.Label(self.summary_frame, text="Promedio de Ejercicios por Tipo: 0.00", font=("Helvetica", 14))
        self.average_label.pack(anchor='w', pady=5)

        self.most_frequent_label = ttk.Label(self.summary_frame, text="Ejercicio Más Frecuente: N/A", font=("Helvetica", 14))
        self.most_frequent_label.pack(anchor='w', pady=5)

        self.total_per_month_label = ttk.Label(self.summary_frame, text="Total de Ejercicios por Mes:", font=("Helvetica", 14))
        self.total_per_month_label.pack(anchor='w', pady=(20,5))

        self.total_per_month_text = tk.Text(self.summary_frame, height=5, width=50, font=("Helvetica", 12),
                                           wrap='word', bg='white')
        self.total_per_month_text.pack(anchor='w', pady=5)
        self.total_per_month_text.configure(state='disabled')

        # Inicializar con datos completos
        self.filtered_history = self.read_history()
        self.update_summary_tab()

    def update_summary_tab(self):
        """Actualiza las estadísticas en la pestaña de resumen."""
        stats = self.process_statistics(self.filtered_history)

        # Actualizar etiquetas
        self.total_label.config(text=f"Total de Ejercicios Resueltos: {stats['total_exercises']}")
        self.average_label.config(text=f"Promedio de Ejercicios por Tipo: {stats['average_per_type']:.2f}")

        # Ejercicio más frecuente
        if stats['exercise_types']:
            most_frequent = max(stats['exercise_types'], key=stats['exercise_types'].get)
            self.most_frequent_label.config(text=f"Ejercicio Más Frecuente: {most_frequent} ({stats['exercise_types'][most_frequent]})")
        else:
            self.most_frequent_label.config(text="Ejercicio Más Frecuente: N/A")

        # Total por mes
        total_per_month = self.calculate_total_per_month(self.filtered_history)
        self.total_per_month_text.configure(state='normal')
        self.total_per_month_text.delete('1.0', tk.END)
        for month, count in sorted(total_per_month.items()):
            self.total_per_month_text.insert(tk.END, f"{month}: {count}\n")
        self.total_per_month_text.configure(state='disabled')

    def calculate_total_per_month(self, history):
        """Calcula el total de ejercicios resueltos por mes."""
        df = pd.DataFrame(history)
        if df.empty:
            return {}
        df['date'] = pd.to_datetime(df['details'].apply(lambda x: x.get('date', '')), errors='coerce')
        df = df.dropna(subset=['date'])
        df['Month'] = df['date'].dt.strftime('%Y-%m')
        total_per_month = df.groupby('Month').size().to_dict()
        return total_per_month

    def init_charts_tab(self):
        """Inicializa los widgets de la pestaña de gráficas."""
        self.charts_frame = ttk.Frame(self.charts_tab)
        self.charts_frame.pack(pady=20, padx=20, fill='both', expand=True)

        # Selección de tipo de gráfica
        chart_selection_frame = ttk.Frame(self.charts_frame)
        chart_selection_frame.pack(pady=10)

        chart_label = ttk.Label(chart_selection_frame, text="Selecciona el tipo de gráfica:", font=("Helvetica", 14))
        chart_label.pack(side='left', padx=(0,10))

        self.chart_type_var = tk.StringVar()
        self.chart_type_combobox = ttk.Combobox(chart_selection_frame, textvariable=self.chart_type_var, state='readonly', style='TCombobox')
        self.chart_type_combobox['values'] = ['Barras', 'Pastel', 'Línea', 'Histograma', 'Dispersión', 'Área']
        self.chart_type_combobox.current(0)
        self.chart_type_combobox.pack(side='left')
        self.chart_type_combobox.bind("<<ComboboxSelected>>", self.update_charts_tab)

        # Frame para la gráfica
        self.single_chart_frame = ttk.Frame(self.charts_frame)
        self.single_chart_frame.pack(fill='both', expand=True, pady=10)

        # Inicializar con datos completos
        self.update_charts_tab()

    def update_charts_tab(self, event=None):
        """Actualiza la gráfica en la pestaña de gráficas según la selección."""
        stats = self.process_statistics(self.filtered_history)

        # Limpiar gráfica existente
        for widget in self.single_chart_frame.winfo_children():
            widget.destroy()

        selected_chart = self.chart_type_var.get()
        if not selected_chart:
            selected_chart = 'Barras'  # Valor por defecto

        if selected_chart == 'Barras':
            self.create_bar_chart(self.single_chart_frame, stats['exercise_types'])
        elif selected_chart == 'Pastel':
            self.create_pie_chart(self.single_chart_frame, stats['exercise_types'])
        elif selected_chart == 'Línea':
            self.create_line_chart(self.single_chart_frame, self.filtered_history)
        elif selected_chart == 'Histograma':
            self.create_histogram(self.single_chart_frame, stats['exercise_types'])
        elif selected_chart == 'Dispersión':
            self.create_scatter_chart(self.single_chart_frame, stats['exercise_types'])
        elif selected_chart == 'Área':
            self.create_area_chart(self.single_chart_frame, stats['exercise_types'])

    def create_bar_chart(self, parent, data):
        """
        Crea un gráfico de barras de los tipos de ejercicios.
        """
        if not data:
            no_data = ttk.Label(parent, text="No hay datos para mostrar.", font=("Helvetica", 12))
            no_data.pack()
            return

        fig = plt.Figure(figsize=(8, 6), dpi=100)
        ax = fig.add_subplot(111)
        types = list(data.keys())
        counts = list(data.values())
        bars = ax.bar(types, counts, color='#3498DB')  # Azul moderno
        ax.set_title('📊 Ejercicios Resueltos por Tipo')
        ax.set_xlabel('Tipo de Ejercicio')
        ax.set_ylabel('Cantidad')
        ax.tick_params(axis='x', rotation=45)

        # Añadir etiquetas a las barras
        for bar in bars:
            height = bar.get_height()
            ax.annotate(f'{height}',
                        xy=(bar.get_x() + bar.get_width() / 2, height),
                        xytext=(0, 3),  # 3 puntos de desplazamiento
                        textcoords="offset points",
                        ha='center', va='bottom')

        fig.tight_layout()

        canvas = FigureCanvasTkAgg(fig, master=parent)
        canvas.draw()
        canvas.get_tk_widget().pack(fill='both', expand=True)

    def create_pie_chart(self, parent, data):
        """
        Crea un gráfico de pastel de los tipos de ejercicios.
        """
        if not data:
            no_data = ttk.Label(parent, text="No hay datos para mostrar.", font=("Helvetica", 12))
            no_data.pack()
            return

        fig = plt.Figure(figsize=(8, 6), dpi=100)
        ax = fig.add_subplot(111)
        types = list(data.keys())
        counts = list(data.values())
        colors_list = plt.cm.Paired(range(len(types)))

        ax.pie(counts, labels=types, autopct='%1.1f%%', startangle=140, colors=colors_list, wedgeprops={'edgecolor': 'black'})
        ax.set_title('📈 Distribución de Ejercicios por Tipo')

        fig.tight_layout()

        canvas = FigureCanvasTkAgg(fig, master=parent)
        canvas.draw()
        canvas.get_tk_widget().pack(fill='both', expand=True)

    def create_line_chart(self, parent, history):
        """
        Crea un gráfico de línea mostrando la evolución de ejercicios resueltos a lo largo del tiempo.
        """
        if not history:
            no_data = ttk.Label(parent, text="No hay datos para mostrar.", font=("Helvetica", 12))
            no_data.pack()
            return

        df = pd.DataFrame(history)
        if df.empty:
            no_data = ttk.Label(parent, text="No hay datos para mostrar.", font=("Helvetica", 12))
            no_data.pack()
            return

        df['date'] = pd.to_datetime(df['details'].apply(lambda x: x.get('date', '')), errors='coerce')
        df = df.dropna(subset=['date'])
        df['Date'] = df['date'].dt.date
        exercises_per_day = df.groupby('Date').size()

        if exercises_per_day.empty:
            no_data = ttk.Label(parent, text="No hay datos para mostrar.", font=("Helvetica", 12))
            no_data.pack()
            return

        fig = plt.Figure(figsize=(8, 6), dpi=100)
        ax = fig.add_subplot(111)
        exercises_per_day.plot(kind='line', ax=ax, color='#E67E22', marker='o')  # Naranja moderno
        ax.set_title('📈 Evolución de Ejercicios Resueltos')
        ax.set_xlabel('Fecha')
        ax.set_ylabel('Cantidad de Ejercicios')
        ax.grid(True)

        fig.tight_layout()

        canvas = FigureCanvasTkAgg(fig, master=parent)
        canvas.draw()
        canvas.get_tk_widget().pack(fill='both', expand=True)

    def create_histogram(self, parent, data):
        """
        Crea un histograma de la frecuencia de ejercicios por tipo.
        """
        if not data:
            no_data = ttk.Label(parent, text="No hay datos para mostrar.", font=("Helvetica", 12))
            no_data.pack()
            return

        fig = plt.Figure(figsize=(8, 6), dpi=100)
        ax = fig.add_subplot(111)
        types = list(data.keys())
        counts = list(data.values())
        ax.bar(types, counts, color='#2ECC71', edgecolor='black')  # Verde moderno
        ax.set_title('📊 Frecuencia de Ejercicios por Tipo')
        ax.set_xlabel('Tipo de Ejercicio')
        ax.set_ylabel('Frecuencia')
        ax.tick_params(axis='x', rotation=45)

        # Añadir etiquetas a las barras
        for i, count in enumerate(counts):
            ax.text(i, count + max(counts)*0.01, str(count), ha='center', va='bottom')

        fig.tight_layout()

        canvas = FigureCanvasTkAgg(fig, master=parent)
        canvas.draw()
        canvas.get_tk_widget().pack(fill='both', expand=True)

    def create_scatter_chart(self, parent, data):
        """
        Crea un gráfico de dispersión de los tipos de ejercicios.
        """
        if not data:
            no_data = ttk.Label(parent, text="No hay datos para mostrar.", font=("Helvetica", 12))
            no_data.pack()
            return

        # Para el gráfico de dispersión, necesitaremos dos variables. Usaremos la frecuencia y un índice.
        fig = plt.Figure(figsize=(8, 6), dpi=100)
        ax = fig.add_subplot(111)

        types = list(data.keys())
        counts = list(data.values())
        indices = list(range(1, len(types) + 1))

        ax.scatter(indices, counts, color='#9B59B6', s=100, alpha=0.7)  # Morado moderno
        ax.set_title('🔍 Dispersión de Ejercicios por Tipo')
        ax.set_xlabel('Tipo de Ejercicio')
        ax.set_ylabel('Frecuencia')
        ax.set_xticks(indices)
        ax.set_xticklabels(types, rotation=45)

        # Añadir etiquetas a los puntos
        for i, count in zip(indices, counts):
            ax.annotate(str(count), (i, count), textcoords="offset points", xytext=(0,10), ha='center')

        fig.tight_layout()

        canvas = FigureCanvasTkAgg(fig, master=parent)
        canvas.draw()
        canvas.get_tk_widget().pack(fill='both', expand=True)

    def create_area_chart(self, parent, data):
        """
        Crea un gráfico de área de los tipos de ejercicios.
        """
        if not data:
            no_data = ttk.Label(parent, text="No hay datos para mostrar.", font=("Helvetica", 12))
            no_data.pack()
            return

        types = list(data.keys())
        counts = list(data.values())

        # Ordenar los datos para un gráfico de área coherente
        types_sorted = sorted(types)
        counts_sorted = [data[t] for t in types_sorted]

        fig = plt.Figure(figsize=(8, 6), dpi=100)
        ax = fig.add_subplot(111)
        ax.fill_between(types_sorted, counts_sorted, color='#34495E', alpha=0.6)  # Azul oscuro moderno
        ax.plot(types_sorted, counts_sorted, color='#34495E', marker='o')
        ax.set_title('📉 Área de Ejercicios Resueltos por Tipo')
        ax.set_xlabel('Tipo de Ejercicio')
        ax.set_ylabel('Frecuencia')
        ax.tick_params(axis='x', rotation=45)

        # Añadir etiquetas a los puntos
        for i, count in enumerate(counts_sorted):
            ax.annotate(str(count), (types_sorted[i], count), textcoords="offset points", xytext=(0,10), ha='center')

        fig.tight_layout()

        canvas = FigureCanvasTkAgg(fig, master=parent)
        canvas.draw()
        canvas.get_tk_widget().pack(fill='both', expand=True)

    def init_export_tab(self):
        """Inicializa los widgets de la pestaña de exportación."""
        export_frame = ttk.Frame(self.export_tab)
        export_frame.pack(pady=50, padx=50)

        # Botón para Exportar a Excel
        export_excel_button = ttk.Button(export_frame, text="💾 Exportar a Excel", command=self.export_to_excel, style='ExportButton.TButton')
        export_excel_button.pack(pady=20, fill='x')

        # Botón para Exportar a PDF
        export_pdf_button = ttk.Button(export_frame, text="📄 Exportar a PDF", command=self.export_to_pdf, style='ExportButton.TButton')
        export_pdf_button.pack(pady=20, fill='x')

    def export_to_excel(self):
        """Exporta las estadísticas filtradas a un archivo Excel con formato de reporte."""
        if not self.filtered_history:
            messagebox.showwarning("Exportar Excel", "No hay datos para exportar.")
            return

        # Solicitar al usuario la ubicación para guardar el Excel
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("Excel files", "*.xlsx")],
                                                 title="Guardar estadísticas como Excel")
        if save_path:
            try:
                # Crear un DataFrame con los datos filtrados
                df = pd.DataFrame(self.filtered_history)
                # Añadir columna de fecha
                df['Fecha'] = pd.to_datetime(df['details'].apply(lambda x: x.get('date', '')), errors='coerce').dt.date

                # Procesar estadísticas
                stats = self.process_statistics(self.filtered_history)
                total_exercises = stats['total_exercises']
                average_per_type = stats['average_per_type']
                exercise_types = stats['exercise_types']
                most_frequent = max(exercise_types, key=stats['exercise_types'].get) if exercise_types else "N/A"
                most_frequent_count = exercise_types[most_frequent] if exercise_types else 0
                total_per_month = self.calculate_total_per_month(self.filtered_history)

                # Crear un ExcelWriter con openpyxl
                with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                    # Escribir los datos filtrados
                    df.to_excel(writer, sheet_name='Datos Filtrados', index=False)

                    # Aplicar formato a la hoja 'Datos Filtrados'
                    workbook = writer.book
                    data_sheet = writer.sheets['Datos Filtrados']

                    # Aplicar estilos a los encabezados
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
                    for col_num, column_title in enumerate(df.columns, 1):
                        cell = data_sheet.cell(row=1, column=col_num)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                        # Auto-ajustar el ancho de las columnas
                        column_letter = get_column_letter(col_num)
                        max_length = max(
                            df[column_title].astype(str).map(len).max(),
                            len(column_title)
                        ) + 2
                        data_sheet.column_dimensions[column_letter].width = max_length

                    # Agregar bordes a todas las celdas
                    thin_border = Border(left=Side(style='thin'),
                                         right=Side(style='thin'),
                                         top=Side(style='thin'),
                                         bottom=Side(style='thin'))

                    for row in data_sheet.iter_rows(min_row=1, max_row=data_sheet.max_row, min_col=1, max_col=data_sheet.max_column):
                        for cell in row:
                            cell.border = thin_border

                    # Crear una hoja para el resumen
                    summary_sheet = workbook.create_sheet(title='Resumen')

                    # Escribir el resumen con estilos
                    summary_sheet['A1'] = "Resumen de Estadísticas"
                    summary_sheet['A1'].font = Font(size=16, bold=True, color="0000FF")
                    summary_sheet['A1'].alignment = Alignment(horizontal="center")

                    summary_sheet.merge_cells('A1:B1')  # Fusionar celdas para el título

                    summary_sheet['A3'] = "Total de Ejercicios Resueltos:"
                    summary_sheet['B3'] = total_exercises

                    summary_sheet['A4'] = "Promedio de Ejercicios por Tipo:"
                    summary_sheet['B4'] = round(average_per_type, 2)

                    summary_sheet['A5'] = "Ejercicio Más Frecuente:"
                    summary_sheet['B5'] = f"{most_frequent} ({most_frequent_count})"

                    summary_sheet['A7'] = "Mes"
                    summary_sheet['B7'] = "Cantidad"

                    row = 8
                    for month, count in sorted(total_per_month.items()):
                        summary_sheet[f'A{row}'] = month
                        summary_sheet[f'B{row}'] = count
                        row += 1

                    # Aplicar estilos a la tabla de resumen
                    for cell in ['A3', 'A4', 'A5', 'A7', 'B7']:
                        summary_sheet[cell].font = Font(bold=True, color="FFFFFF")
                        summary_sheet[cell].fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
                        summary_sheet[cell].alignment = Alignment(horizontal="center")
                        summary_sheet[cell].border = thin_border

                    for r in range(3, row):
                        for c in ['A', 'B']:
                            cell = summary_sheet[f'{c}{r}']
                            if c == 'A' and r >= 8:
                                cell.font = Font(bold=False, color="000000")
                                cell.fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
                            elif c == 'B' and r >= 8:
                                cell.font = Font(bold=False, color="000000")
                                cell.fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal="center")

                    # Auto-ajustar el ancho de las columnas en la hoja de resumen
                    for col_num, column_title in enumerate(['A', 'B'], 1):
                        column_letter = get_column_letter(col_num)
                        max_length = max(
                            [len(str(cell.value)) for cell in summary_sheet[column_letter]] + [len(column_title)]
                        ) + 2
                        summary_sheet.column_dimensions[column_letter].width = max_length

            except Exception as e:
                messagebox.showerror("Error Exportar Excel", f"No se pudo generar el archivo Excel.\n{e}")
            else:
                messagebox.showinfo("Exportar Excel", f"Estadísticas exportadas exitosamente en:\n{save_path}")

    def export_to_pdf(self):
        """Exporta las estadísticas filtradas a un archivo PDF."""
        if not self.filtered_history:
            messagebox.showwarning("Exportar PDF", "No hay datos para exportar.")
            return

        # Solicitar al usuario la ubicación para guardar el PDF
        save_path = filedialog.asksaveasfilename(defaultextension=".pdf",
                                                 filetypes=[("PDF files", "*.pdf")],
                                                 title="Guardar estadísticas como PDF")
        if save_path:
            try:
                # Procesar estadísticas
                stats = self.process_statistics(self.filtered_history)
                total_exercises = stats['total_exercises']
                average_per_type = stats['average_per_type']
                exercise_types = stats['exercise_types']
                most_frequent = max(exercise_types, key=stats['exercise_types'].get) if exercise_types else "N/A"
                most_frequent_count = exercise_types[most_frequent] if exercise_types else 0
                total_per_month = self.calculate_total_per_month(self.filtered_history)

                # Crear PDF
                c = canvas.Canvas(save_path, pagesize=letter)
                width, height = letter
                styles = getSampleStyleSheet()
                style_heading = styles['Heading1']
                style_normal = styles['Normal']

                # Título
                c.setFont("Helvetica-Bold", 20)
                c.drawString(50, height - 50, "Estadísticas de Uso - Algebrify")

                # Resumen
                y = height - 80
                c.setFont("Helvetica-Bold", 14)
                c.drawString(50, y, "Resumen de Estadísticas")
                y -= 30

                c.setFont("Helvetica", 12)
                c.drawString(50, y, f"Total de Ejercicios Resueltos: {total_exercises}")
                y -= 20
                c.drawString(50, y, f"Promedio de Ejercicios por Tipo: {average_per_type:.2f}")
                y -= 20
                c.drawString(50, y, f"Ejercicio Más Frecuente: {most_frequent} ({most_frequent_count})")
                y -= 30

                # Tabla de ejercicios por mes
                c.setFont("Helvetica-Bold", 14)
                c.drawString(50, y, "Total de Ejercicios por Mes:")
                y -= 20

                data = [["Mes", "Cantidad"]]
                for month, count in sorted(total_per_month.items()):
                    data.append([month, count])

                table = Table(data, colWidths=[200, 100])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#3498DB")),
                    ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                    ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                    ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0,0), (-1,0), 12),
                    ('BOTTOMPADDING', (0,0), (-1,0), 12),
                    ('BACKGROUND', (0,1), (-1,-1), colors.HexColor("#ECF0F1")),
                    ('GRID', (0,0), (-1,-1), 1, colors.black),
                ]))
                table.wrapOn(c, width, height)
                table.drawOn(c, 50, y - 20*len(data))

                c.save()
                messagebox.showinfo("Exportar PDF", f"Estadísticas exportadas exitosamente en:\n{save_path}")
            except Exception as e:
                messagebox.showerror("Error Exportar PDF", f"No se pudo generar el archivo PDF.\n{e}")

    def read_history(self):
        if os.path.exists(HISTORY_FILE):
            with open(HISTORY_FILE, 'r') as f:
                try:
                    history = json.load(f)
                except json.JSONDecodeError:
                    history = []
        else:
            history = []
        return history

    def process_statistics(self, history):
        """
        Procesa los datos del historial para generar estadísticas.
        """
        exercise_types = {}
        total_exercises = len(history)

        for exercise in history:
            exercise_type = exercise.get('type', 'Desconocido')
            exercise_types[exercise_type] = exercise_types.get(exercise_type, 0) + 1

        # Calcular el promedio de ejercicios por tipo
        average = total_exercises / len(exercise_types) if exercise_types else 0

        return {
            'total_exercises': total_exercises,
            'exercise_types': exercise_types,
            'average_per_type': average
        }

    def create_stats_summary(self, parent, stats):
        """
        Crea un resumen de las estadísticas clave.
        """
        summary_frame = ttk.Frame(parent)
        summary_frame.pack(pady=10)

        # Total de ejercicios
        total_label = ttk.Label(summary_frame,
                               text=f"Total de Ejercicios Resueltos: {stats['total_exercises']}",
                               font=("Helvetica", 14))
        total_label.pack(anchor='w', pady=5)

        # Promedio de ejercicios por tipo
        average_label = ttk.Label(summary_frame,
                                 text=f"Promedio de Ejercicios por Tipo: {stats['average_per_type']:.2f}",
                                 font=("Helvetica", 14))
        average_label.pack(anchor='w', pady=5)

        # Ejercicio más frecuente
        if stats['exercise_types']:
            most_frequent = max(stats['exercise_types'], key=stats['exercise_types'].get)
            most_frequent_count = stats['exercise_types'][most_frequent]
            most_frequent_label = ttk.Label(summary_frame,
                                           text=f"Ejercicio Más Frecuente: {most_frequent} ({most_frequent_count})",
                                           font=("Helvetica", 14))
            most_frequent_label.pack(anchor='w', pady=5)
        else:
            most_frequent_label = ttk.Label(summary_frame,
                                           text="Ejercicio Más Frecuente: N/A",
                                           font=("Helvetica", 14))
            most_frequent_label.pack(anchor='w', pady=5)

def main():
    # Este método es solo para pruebas independientes de Estadisticas.py
    root = tk.Tk()
    root.geometry("1000x700")
    # Definir colores si parent no está definido
    class Parent:
        bg_color = 'white'
        fg_color = 'black'
    app = Estadisticas(root, parent=Parent())
    root.mainloop()

if __name__ == "__main__":
    main()
